import os
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from backend.config import EXPORT_DIR

def _ensure_export_dir():
    if not os.path.exists(EXPORT_DIR):
        os.makedirs(EXPORT_DIR)

def export_archive_csv(archive_data: list[dict], output_filename: str) -> str:
    """Export the archive as a basic CSV file."""
    _ensure_export_dir()
    filepath = os.path.join(EXPORT_DIR, output_filename)
    
    if not archive_data:
        # Create empty file
        open(filepath, 'w').close()
        return filepath
        
    keys = list(archive_data[0].keys())
    
    with open(filepath, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=keys)
        writer.writeheader()
        for row in archive_data:
            writer.writerow(row)
            
    return filepath

def export_archive_excel(archive_data: list[dict], output_filename: str) -> str:
    """Export the archive into a formatted Excel file with multiple sheets."""
    _ensure_export_dir()
    filepath = os.path.join(EXPORT_DIR, output_filename)
    
    wb = Workbook()
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F8BF9")
    align_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    # Helper to create sheet with formatted headers
    def create_sheet_with_headers(ws, title, headers):
        ws.title = title
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            ws.column_dimensions[cell.column_letter].width = 25
            
    # Sheet 1: Résumé
    ws_summary = wb.active
    create_sheet_with_headers(ws_summary, "Résumé Général", 
                              ["ID", "Date", "Type", "Niveau", "Thème", "Titre", "Favori"])
    
    for item in archive_data:
        ws_summary.append([
            item.get('id', ''),
            str(item.get('date_created', '')),
            item.get('content_type', ''),
            item.get('niveau', ''),
            item.get('theme', ''),
            item.get('title', ''),
            "Oui" if item.get('is_favorite') else "Non"
        ])
    
    # Sheet 2: Textes
    textes = [item for item in archive_data if item.get('content_type') == 'Texte']
    if textes:
        ws_textes = wb.create_sheet("Textes de support")
        create_sheet_with_headers(ws_textes, "Textes de support", 
                                  ["Titre", "Niveau", "Thème", "Contenu", "Notes Pédagogiques"])
        
        for item in textes:
            try:
                import json
                content_dict = json.loads(item.get('content_json', '{}'))
                ws_textes.append([
                    item.get('title', ''),
                    item.get('niveau', ''),
                    item.get('theme', ''),
                    content_dict.get('texte', ''),
                    content_dict.get('notes_pedagogiques', '')
                ])
                # Set alignment for text columns
                ws_textes.cell(row=ws_textes.max_row, column=4).alignment = align_left
                ws_textes.cell(row=ws_textes.max_row, column=5).alignment = align_left
            except Exception as exc:
                import logging
                logging.getLogger(__name__).warning("Erreur lors de l'export Excel (Texte ignoré): %s", exc)
                pass

    # Similar sheets could be added for Situations and Grilles
    wb.save(filepath)
    return filepath

def export_grille_excel(grille_data: dict, output_filename: str) -> str:
    """Export a rubric to an Excel file with color coding."""
    _ensure_export_dir()
    filepath = os.path.join(EXPORT_DIR, output_filename)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Grille Evaluation"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F8BF9")
    align_wrap = Alignment(wrap_text=True, vertical="top")
    
    # Headers
    headers = ["Critère", "Très Satisfaisant", "Satisfaisant", "Peu Satisfaisant", "Insuffisant"]
    ws.append(headers)
    
    for col in range(1, 6):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        ws.column_dimensions[cell.column_letter].width = 30
        
    criteres = grille_data.get('criteres', [])
    for c in criteres:
        nom = c.get('nom', '')
        niveaux = c.get('niveaux', {})
        
        row_data = [
            f"{nom}\n{c.get('description', '')}",
            f"{niveaux.get('excellent', {}).get('description', '')}\n({niveaux.get('excellent', {}).get('points', '')} pts)",
            f"{niveaux.get('bien', {}).get('description', '')}\n({niveaux.get('bien', {}).get('points', '')} pts)",
            f"{niveaux.get('passable', {}).get('description', '')}\n({niveaux.get('passable', {}).get('points', '')} pts)",
            f"{niveaux.get('insuffisant', {}).get('description', '')}\n({niveaux.get('insuffisant', {}).get('points', '')} pts)"
        ]
        ws.append(row_data)
        
        for col in range(1, 6):
            ws.cell(row=ws.max_row, column=col).alignment = align_wrap
            
    wb.save(filepath)
    return filepath
